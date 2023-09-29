'''Train CIFAR10 with PyTorch.'''
import torch
import torch.nn as nn
import torch.optim as optim
import torch.nn.functional as F
import torch.backends.cudnn as cudnn
from torch.optim.lr_scheduler import _LRScheduler
import time
import datetime as dt

import torchvision
import torchvision.transforms as transforms
# from torchstat import stat

import os
import argparse

from tensorboardX import SummaryWriter

from models import *
from datetime import datetime

import random

from torchsummary import summary
# Training

parser = argparse.ArgumentParser(description='PyTorch CIFAR10 Training')
parser.add_argument('--tao', default=1, type=int)
parser.add_argument('--nnlayer', default=1, type=int)
parser.add_argument('--repnum', default=1, type=int)
parser.add_argument('--h', default=1.0, type=float)
parser.add_argument('--b1', default=1.5, type=float)
parser.add_argument('--b2', default=-0.5, type=float)
parser.add_argument('--b3', default=1.0, type=float)
parser.add_argument('--a0', default=1.0, type=float)
parser.add_argument('--lr', default=0.1, type=float, help='learning rate')
parser.add_argument('--resume', '-r', action='store_true',help='resume from checkpoint')
parser.add_argument('--epoch', type=int, default=200, help='training epoch')
parser.add_argument('--warm', type=int, default=1, help='warm up training phase')

                                    
parser.add_argument('-d', '--data', default='./data', type=str)
parser.add_argument('--arch', '-a', default='ResNet18', type=str)
parser.add_argument('-j', '--workers', default=8, type=int, metavar='N',
                    help='number of data loading workers (default: 4)')

parser.add_argument('--start-epoch', default=0, type=int, metavar='N',
                    help='manual epoch number (useful on restarts)')
parser.add_argument('-bs', '--batch_size', default=256, type=int,
                    metavar='N', help='mini-batch size (default: 256)')
parser.add_argument('--test-batch', default=32, type=int, metavar='N',
                    help='test batchsize (default: 200)')

parser.add_argument('--momentum', default=0.9, type=float, metavar='M',
                    help='momentum')
parser.add_argument('--print-freq', '-p', default=250, type=int,
                    metavar='N', help='print frequency (default: 10)')
parser.add_argument('-c', '--checkpoint', type=str, metavar='PATH',
                    help='path to save checkpoint (default: checkpoint)')

parser.add_argument('-e', '--evaluate', dest='evaluate', action='store_true',
                    help='evaluate model on validation set')
parser.add_argument('--pretrained', dest='pretrained', action='store_true',
                    help='use pre-trained model')
# Optimization options
parser.add_argument('--opt_level', default='O2', type=str,
                    help='O2 is fast mixed FP16/32 training, O0 (FP32 training) and O3 (FP16 training), O1 ("conservative mixed precision"), O2 ("fast mixed precision").--opt_level O1 and O2 both use dynamic loss scaling by default unless manually overridden. --opt-level O0 and O3 (the "pure" training modes) do not use loss scaling by default. See more in https://github.com/NVIDIA/apex/tree/f5cd5ae937f168c763985f627bbf850648ea5f3f/examples/imagenet')
parser.add_argument('--keep-batchnorm-fp32', default=True, action='store_true',
                    help='keeping cudnn bn leads to fast training')
parser.add_argument('--loss-scale', type=float, default=None)
parser.add_argument('--dali_cpu', action='store_true',
                    help='Runs CPU based version of DALI pipeline.')
parser.add_argument('--prof', dest='prof', action='store_true',
                    help='Only run 10 iterations for profiling.')
parser.add_argument('-t', '--test', action='store_true',
                    help='Launch test mode with preset arguments')
parser.add_argument('--warmup', '--wp', default=5, type=int,
                    help='number of epochs to warmup')
parser.add_argument('--weight-decay', '--wd', default=4e-5, type=float,
                    metavar='W', help='weight decay (default: 4e-5 for mobile models)')
parser.add_argument('--wd-all', dest = 'wdall', action='store_true',
                    help='weight decay on all parameters')

parser.add_argument("--local_rank", default=0, type=int)
parser.add_argument("--ex", default=0, type=int)
parser.add_argument("--alpha", default=0.1, type=float)
parser.add_argument("--beta", default=15.0, type=float)
parser.add_argument("--notes", default='', type=str)

parser.add_argument("--table_loc", default='', type=str)

args = parser.parse_args()
args.save_path = 'runs/cifar10/'+args.arch+'/Ori'+\
                                    '_BS'+str(args.batch_size)+'LR'+\
                                   str(args.lr)+'epoch'+\
                                   str(args.epoch)+'warmup'+str(args.warm)+\
                                   args.notes+\
                                    "{0:%Y-%m-%dT%H-%M/}".format(datetime.now())

def train(epoch):
    batch_time = AverageMeter('Time', ':6.3f')
    losses = AverageMeter('Loss', ':.4e')
    top1 = AverageMeter('Acc@1', ':6.2f')
    top5 = AverageMeter('Acc@5', ':6.2f')
    end = time.time()
    net.train()

    for batch_idx, (inputs, targets) in enumerate(trainloader):
        # if epoch <= args.warm:
            # warmup_scheduler.step()

        inputs, targets = inputs.to(device), targets.to(device)
        optimizer.zero_grad()
        outputs = net(inputs)
        loss = criterion(outputs, targets)
        acc1, acc5 = accuracy(outputs, targets, topk=(1, 5))
        losses.update(loss.item(), inputs.size(0))
        top1.update(acc1[0], inputs.size(0))
        top5.update(acc5[0], inputs.size(0))

        loss.backward()
        optimizer.step()

        batch_time.update(time.time() - end)
        end = time.time()

    print('Epoch: {:.1f}, Train set: Average loss: {:.4f}, Accuracy: {:.4f}'.format(epoch, losses.avg, top1.avg))
    writer.add_scalar('Train/Average loss', losses.avg, epoch)
    writer.add_scalar('Train/Accuracy-top1', top1.avg, epoch)
    writer.add_scalar('Train/Accuracy-top5', top5.avg, epoch)
    writer.add_scalar('Train/Time', batch_time.sum, epoch)
    #for name, param in net.named_parameters():
        #layer, attr = os.path.splitext(name)
        #attr = attr[1:]
        #writer.add_histogram("{}/{}".format(layer, attr), param, epoch)
    return top1.avg, losses.avg, batch_time.sum


def test(epoch):
    batch_time = AverageMeter('Time', ':6.3f')
    losses = AverageMeter('Loss', ':.4e')
    top1 = AverageMeter('Acc@1', ':6.2f')
    top5 = AverageMeter('Acc@5', ':6.2f')
    end = time.time()
    net.eval()

    with torch.no_grad():
        for batch_idx, (inputs, targets) in enumerate(testloader):
            inputs, targets = inputs.to(device), targets.to(device)
            outputs = net(inputs)
            loss = criterion(outputs, targets)
            acc1, acc5 = accuracy(outputs, targets, topk=(1, 5))
            losses.update(loss.item(), inputs.size(0))
            top1.update(acc1[0], inputs.size(0))
            top5.update(acc5[0], inputs.size(0))
            batch_time.update(time.time() - end)
            end = time.time()
            # test_loss += loss.item()
            # _, preds = outputs.max(1)
            # correct += preds.eq(old_labels).sum()
    print('Test set: Average loss: {:.4f}, Accuracy: {:.4f}'.format(losses.avg, top1.avg))

    #add informations to tensorboard
    writer.add_scalar('Test/Average loss', losses.avg, epoch)
    writer.add_scalar('Test/Accuracy-top1', top1.avg, epoch)
    writer.add_scalar('Test/Accuracy-top5', top5.avg, epoch)
    writer.add_scalar('Test/Time', batch_time.sum, epoch)
    
    return top1.avg, losses.avg, batch_time.sum
            

def Normalize(tensor, mean, std, inplace=False):
    """Normalize a float tensor image with mean and standard deviation.
    This transform does not support PIL Image.

    .. note::
        This transform acts out of place by default, i.e., it does not mutates the input tensor.

    See :class:`~torchvision.transforms.Normalize` for more details.

    Args:
        tensor (Tensor): Float tensor image of size (C, H, W) or (B, C, H, W) to be normalized.
        mean (sequence): Sequence of means for each channel.
        std (sequence): Sequence of standard deviations for each channel.
        inplace(bool,optional): Bool to make this operation inplace.

    Returns:
        Tensor: Normalized Tensor image.
    """
    if not isinstance(tensor, torch.Tensor):
        raise TypeError(
            'Input tensor should be a torch tensor. Got {}.'.format(type(tensor)))

    if not tensor.is_floating_point():
        raise TypeError(
            'Input tensor should be a float tensor. Got {}.'.format(tensor.dtype))

    if tensor.ndim < 3:
        raise ValueError('Expected tensor to be a tensor image of size (..., C, H, W). Got tensor.size() = '
                         '{}.'.format(tensor.size()))

    if not inplace:
        tensor = tensor.clone()

    dtype = tensor.dtype
    mean = torch.as_tensor(mean, dtype=dtype, device=tensor.device)
    std = torch.as_tensor(std, dtype=dtype, device=tensor.device)
    if (std == 0).any():
        raise ValueError(
            'std evaluated to zero after conversion to {}, leading to division by zero.'.format(dtype))
    if mean.ndim == 1:
        mean = mean.view(-1, 1, 1)
    if std.ndim == 1:
        std = std.view(-1, 1, 1)
    tensor.sub_(mean).div_(std)
    return tensor


def UnNormalize(tensor, mean, std, inplace=False):
    """Normalize a float tensor image with mean and standard deviation.
    This transform does not support PIL Image.

    .. note::
        This transform acts out of place by default, i.e., it does not mutates the input tensor.

    See :class:`~torchvision.transforms.Normalize` for more details.

    Args:
        tensor (Tensor): Float tensor image of size (C, H, W) or (B, C, H, W) to be normalized.
        mean (sequence): Sequence of means for each channel.
        std (sequence): Sequence of standard deviations for each channel.
        inplace(bool,optional): Bool to make this operation inplace.

    Returns:
        Tensor: Normalized Tensor image.
    """
    if not isinstance(tensor, torch.Tensor):
        raise TypeError(
            'Input tensor should be a torch tensor. Got {}.'.format(type(tensor)))

    if not tensor.is_floating_point():
        raise TypeError(
            'Input tensor should be a float tensor. Got {}.'.format(tensor.dtype))

    if tensor.ndim < 3:
        raise ValueError('Expected tensor to be a tensor image of size (..., C, H, W). Got tensor.size() = '
                         '{}.'.format(tensor.size()))

    if not inplace:
        tensor = tensor.clone()

    dtype = tensor.dtype
    mean = torch.as_tensor(mean, dtype=dtype, device=tensor.device)
    std = torch.as_tensor(std, dtype=dtype, device=tensor.device)
    if (std == 0).any():
        raise ValueError(
            'std evaluated to zero after conversion to {}, leading to division by zero.'.format(dtype))
    if mean.ndim == 1:
        mean = mean.view(-1, 1, 1)
    if std.ndim == 1:
        std = std.view(-1, 1, 1)
    tensor.mul_(std).add_(mean)
    # tensor.sub_(mean).div_(std)
    return tensor
          
def test_noise(model, noise_type, noise_coff):
    batch_time = AverageMeter('Time', ':6.3f')
    losses = AverageMeter('Loss', ':.4e')
    top1 = AverageMeter('Acc@1', ':6.2f')
    top5 = AverageMeter('Acc@5', ':6.2f')
    end = time.time()
    model.eval()

    with torch.no_grad():
        for batch_idx, (inputs, targets) in enumerate(testloader):
            inputs_pertur, targets = inputs.to(device), targets.to(device)
            
            # inputs_pertur = inputs_pertur + noise_coff * \
            #     torch.autograd.Variable(torch.randn(
            #         inputs_pertur.size()).cuda(), requires_grad=False)
            # # 截断
            # inputs_pertur = F.relu(
            #     F.relu(inputs_pertur.mul_(-1).add_(1)).mul_(-1).add_(1))

            inputs_pertur = UnNormalize(
                inputs_pertur, (0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))

            if noise_type == 'randn':
                inputs_pertur = inputs_pertur + noise_coff * torch.randn_like(inputs_pertur)

            elif noise_type == 'rand':
                inputs_pertur = inputs_pertur + noise_coff * torch.rand_like(inputs_pertur)

            elif noise_type == 'const':
                inputs_pertur = inputs_pertur + noise_coff

            inputs_pertur = torch.clamp(inputs_pertur, 0, 1)
            
            inputs_pertur = Normalize(
                inputs_pertur, (0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))
            
            outputs = model(inputs_pertur)
            loss = criterion(outputs, targets)
            acc1, acc5 = accuracy(outputs, targets, topk=(1, 5))
            losses.update(loss.item(), inputs.size(0))
            top1.update(acc1[0], inputs.size(0))
            top5.update(acc5[0], inputs.size(0))
            batch_time.update(time.time() - end)
            end = time.time()
            # test_loss += loss.item()
            # _, preds = outputs.max(1)
            # correct += preds.eq(old_labels).sum()
    # print('ADD noise Test set: Average loss: {:.4f}, Accuracy: {:.4f}'.format(losses.avg, top1.avg))



    
    return top1.avg, losses.avg, batch_time.sum



class AverageMeter(object):
    """Computes and stores the average and current value"""
    def __init__(self, name, fmt=':f'):
        self.name = name
        self.fmt = fmt
        self.reset()

    def reset(self):
        self.val = 0
        self.avg = 0
        self.sum = 0
        self.count = 0

    def update(self, val, n=1):
        self.val = val
        self.sum += val * n
        self.count += n
        self.avg = self.sum / self.count

    def __str__(self):
        fmtstr = '{name} {val' + self.fmt + '} ({avg' + self.fmt + '})'
        return fmtstr.format(**self.__dict__)


def accuracy(output, target, topk=(1,)):
    """Computes the accuracy over the k top predictions for the specified values of k"""
    with torch.no_grad():
        maxk = max(topk)
        batch_size = target.size(0)

        _, pred = output.topk(maxk, 1, True, True)
        pred = pred.t()
        correct = pred.eq(target.view(1, -1).expand_as(pred))

        res = []
        for k in topk:
            #correct_k = correct[:k].view(-1).float().sum(0, keepdim=True)
            correct_k = correct[:k].contiguous().view(-1).float().sum(0, keepdim=True)
            res.append(correct_k.mul_(100.0 / batch_size))
        return res

class WarmUpLR(_LRScheduler):
    """warmup_training learning rate scheduler
    Args:
        optimizer: optimzier(e.g. SGD)
        total_iters: totoal_iters of warmup phase
    """
    def __init__(self, optimizer, total_iters, last_epoch=-1):
        
        self.total_iters = total_iters
        super().__init__(optimizer, last_epoch)

    def get_lr(self):
        """we will use the first m batches, and set the learning
        rate to base_lr * m / total_iters
        """
        return [base_lr * self.last_epoch / (self.total_iters + 1e-8) for base_lr in self.base_lrs]
        
        
if __name__ == '__main__':
    
    seed=args.repnum
    def seed_torch(seed):
        random.seed(seed)
        os.environ['PYTHONHASHSEED'] = str(seed) 
        np.random.seed(seed)
        torch.manual_seed(seed)
        torch.cuda.manual_seed(seed)
        torch.cuda.manual_seed_all(seed) # if you are using multi-GPU.
        torch.backends.cudnn.benchmark = True
        torch.backends.cudnn.deterministic = True
    seed_torch(seed)
    print("种子值为",seed)
    
    

  

    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    best_acc = 0  # best test accuracy
    start_epoch = 0  # start from epoch 0 or last checkpoint epoch
    #os.environ['CUDA_VISIBLE_DEVICES'] = '2,3'
    # Data
    print('==> Preparing data..')
    transform_train = transforms.Compose([
        transforms.RandomCrop(32, padding=4),
        transforms.RandomHorizontalFlip(),
        transforms.ToTensor(),
        transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010)),
    ])

    transform_test = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010)),
        
    ])

    trainset = torchvision.datasets.CIFAR10(
        root='./data', train=True, download=True, transform=transform_train)
    trainloader = torch.utils.data.DataLoader(
        trainset, args.batch_size, shuffle=False, num_workers=args.workers)
        # trainset, args.batch_size, shuffle=True, num_workers=args.workers)

    testset = torchvision.datasets.CIFAR10(
        root='./data', train=False, download=True, transform=transform_test)
    testloader = torch.utils.data.DataLoader(
        testset, args.batch_size, shuffle=False, num_workers=args.workers)

    classes = ('plane', 'car', 'bird', 'cat', 'deer',
               'dog', 'frog', 'horse', 'ship', 'truck')
    
    # Model
    
    
    print('==> Building model {}'.format( args.arch ))
    print('==> b1 {}'.format( args.b1 ))
    print('==> b2 {}'.format( args.b2 ))
    print('==> b3 {}'.format( args.b3 ))
    print('==> a0 {}'.format( args.a0 ))
    print('==> Train info:')
    print('==> Start DateTime: {}'.format( dt.datetime.now() ))
    print('==> Device: {}'.format( args.notes ))
    net_name = args.arch
    model_name = args.arch
    if args.arch == "r18" :
        net = ResNet18()
    elif args.arch == "r18v2" :
        net = ResNet18V2()
    elif args.arch=="r34" :
        net = ResNet34()
    elif args.arch=="r34v2" :
        net = ResNet34V2()
    elif args.arch=="r50" :
        net = ResNet50()
    elif args.arch=="r50v2" :
        net = ResNet50V2()
    elif args.arch=="r101" :
        net = ResNet101()
    elif args.arch=="r152" :
        net = ResNet152()

    elif args.arch=="pr18" :
        net = PreActResNet18()
    elif args.arch=="pr34" :
        net = PreActResNet34()
    elif args.arch=="pr50" :
        net = PreActResNet50()
    elif args.arch=="pr101" :
        net = PreActResNet101()
    elif args.arch=="pr152" :
        net = PreActResNet152()
        
    elif args.arch=="ab2r18" :
        net = AB2_ResNet18()
    elif args.arch=="ab2r34" :
        net = AB2_ResNet34()
    elif args.arch=="ab2r50" :
        net = AB2_ResNet50()
    elif args.arch=="ab2r101" :
        net = AB2_ResNet101()
    elif args.arch=="ab2r152" :
        net = AB2_ResNet152()    
    elif args.arch=="ab2pr152" :
        net = AB2_PreActResNet152()
    elif args.arch=="ab2pr101" :
        net = AB2_PreActResNet101()
    elif args.arch=="ab2pr50" :
        net = AB2_PreActResNet50()
    elif args.arch=="ab2pr34" :
        net = AB2_PreActResNet34()
    elif args.arch=="ab2pr18" :
        net = AB2_PreActResNet18()
        
    elif args.arch=="ab3r18" :
        net = AB3_ResNet18()
    elif args.arch=="ab3r34" :
        net = AB3_ResNet34()
    elif args.arch=="ab3r50" :
        net = AB3_ResNet50()
    elif args.arch=="ab3r101" :
        net = AB3_ResNet101()
    elif args.arch=="ab3r152" :
        net = AB3_ResNet152()
        
        
    
    elif args.arch=="ab2pr20" :
        net = ab2preresnet(depth=20, k1=args.b1, k2=args.b2, h=args.h, num_classes=10)
        
    elif args.arch=="ab2pr44" :
        net = ab2preresnet(depth=44, k1=args.b1, k2=args.b2, h=args.h, num_classes=10)
        
    elif args.arch=="ab2pr56" :
        net = ab2preresnet(depth=56, k1=args.b1, k2=args.b2, h=args.h, num_classes=10)
        
    elif args.arch=="ab2pr110" :
        net = ab2preresnet(depth=110, k1=args.b1, k2=args.b2, h=args.h, num_classes=10)
        
    elif args.arch=="pr20" :
        net = preresnet20(depth=20, num_classes=10)
    
    elif args.arch=="ab2r20" :
        net = ab2resnet20(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r32" :
        net = ab2resnet32(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r44" :
        net = ab2resnet44(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r56" :
        net = ab2resnet56(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r68" :
        net = ab2resnet68(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r80" :
        net = ab2resnet80(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r92" :
        net = ab2resnet92(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r104" :
        net = ab2resnet104(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r110" :
        net = ab2resnet110(h=args.h, b1=args.b1, b2=args.b2)
        
    elif args.arch=="ab2r20_a" :
        net = ab2resnet20_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r32_a" :
        net = ab2resnet32_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r44_a" :
        net = ab2resnet44_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r56_a" :
        net = ab2resnet56_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r68_a" :
        net = ab2resnet68_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r80_a" :
        net = ab2resnet80_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r92_a" :
        net = ab2resnet92_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r104_a" :
        net = ab2resnet104_a(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r110_a" :
        net = ab2resnet110_a(h=args.h, b1=args.b1, b2=args.b2)
        
    elif args.arch=="ab2r20_b" :
        net = ab2resnet20_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r32_b" :
        net = ab2resnet32_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r44_b" :
        net = ab2resnet44_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r56_b" :
        net = ab2resnet56_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r68_b" :
        net = ab2resnet68_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r80_b" :
        net = ab2resnet80_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r92_b" :
        net = ab2resnet92_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r104_b" :
        net = ab2resnet104_b(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r110_b" :
        net = ab2resnet110_b(h=args.h, b1=args.b1, b2=args.b2)
        
    elif args.arch=="ab2r20_c" :
        net = ab2resnet20_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r32_c" :
        net = ab2resnet32_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r44_c" :
        net = ab2resnet44_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r56_c" :
        net = ab2resnet56_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r68_c" :
        net = ab2resnet68_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r80_c" :
        net = ab2resnet80_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r92_c" :
        net = ab2resnet92_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r104_c" :
        net = ab2resnet104_c(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r110_c" :
        net = ab2resnet110_c(h=args.h, b1=args.b1, b2=args.b2)
        
    elif args.arch=="ab2r20_f" :
        net = ab2resnet20_f(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r110_f" :
        net = ab2resnet110_f(h=args.h, b1=args.b1, b2=args.b2)
        
    elif args.arch=="ab2r20_f1" :
        net = ab2resnet20_f1(h=args.h, b1=args.b1, b2=args.b2)
    elif args.arch=="ab2r110_f1" :
        net = ab2resnet110_f1(h=args.h, b1=args.b1, b2=args.b2)
    
    elif args.arch=="ab3r20_f" :
        net = ab3resnet20_f(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r110_f" :
        net = ab3resnet110_f(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
        
    elif args.arch=="ab3r20_f1" :
        net = ab3resnet20_f1(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r110_f1" :
        net = ab3resnet110_f1(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)

    
    elif args.arch=="ab3r68" :
        net = ab3resnet68(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r80" :
        net = ab3resnet80(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r92" :
        net = ab3resnet92(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r104" :
        net = ab3resnet104(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    
    elif args.arch=="ab3r20" :
        net = ab3resnet20(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r26" :
        net = ab3resnet26(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r32" :
        net = ab3resnet32(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r38" :
        net = ab3resnet38(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r44" :
        net = ab3resnet44(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r50a" :
        net = ab3resnet50a(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r56" :
        net = ab3resnet56(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r62" :
        net = ab3resnet62(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r68" :
        net = ab3resnet68(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r74" :
        net = ab3resnet74(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r80" :
        net = ab3resnet80(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r86" :
        net = ab3resnet86(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r92" :
        net = ab3resnet92(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r98" :
        net = ab3resnet98(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r104" :
        net = ab3resnet104(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r110" :
        net = ab3resnet110(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r116" :
        net = ab3resnet116(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r122" :
        net = ab3resnet122(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r128" :
        net = ab3resnet128(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r134" :
        net = ab3resnet134(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r140" :
        net = ab3resnet140(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r146" :
        net = ab3resnet146(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r152a" :
        net = ab3resnet152a(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    elif args.arch=="ab3r158" :
        net = ab3resnet158(h=args.h, b1=args.b1, b2=args.b2, b3=args.b3)
    
    elif args.arch=="ab1r20" :
        net = ab1resnet20(h=args.h,b1=args.b1)
    elif args.arch=="ab1r32" :
        net = ab1resnet32(h=args.h,b1=args.b1)
    elif args.arch=="ab1r44" :
        net = ab1resnet44(h=args.h,b1=args.b1)
    elif args.arch=="ab1r56" :
        net = ab1resnet56(h=args.h,b1=args.b1)
    elif args.arch=="ab1r98" :
        net = ab1resnet98(h=args.h,b1=args.b1)
    elif args.arch=="ab1r104" :
        net = ab1resnet104(h=args.h,b1=args.b1)
    elif args.arch=="ab1r110" :
        net = ab1resnet110(h=args.h,b1=args.b1)
    elif args.arch=="ab1r116" :
        net = ab1resnet116(h=args.h,b1=args.b1)
    elif args.arch=="ab1r122" :
        net = ab1resnet122(h=args.h,b1=args.b1)
    elif args.arch=="ab1r128" :
        net = ab1resnet128(h=args.h,b1=args.b1)
    elif args.arch=="ab1r134" :
        net = ab1resnet134(h=args.h,b1=args.b1)
    elif args.arch=="ab1r140" :
        net = ab1resnet140(h=args.h,b1=args.b1)
    elif args.arch=="ab1r146" :
        net = ab1resnet146(h=args.h,b1=args.b1)        
    elif args.arch=="ab1r152" :
        net = ab1resnet152(h=args.h,b1=args.b1)
    elif args.arch=="ab1r158" :
        net = ab1resnet158(h=args.h,b1=args.b1)
    elif args.arch=="ab1r164" :
        net = ab1resnet164(h=args.h,b1=args.b1)
    elif args.arch=="ab1r170" :
        net = ab1resnet170(h=args.h,b1=args.b1)
    elif args.arch=="ab1r176" :
        net = ab1resnet176(h=args.h,b1=args.b1)
    elif args.arch=="ab1r182" :
        net = ab1resnet182(h=args.h,b1=args.b1)
    elif args.arch=="ab1r188" :
        net = ab1resnet188(h=args.h,b1=args.b1)
    elif args.arch=="ab1r218" :
        net = ab1resnet218(h=args.h,b1=args.b1)
    elif args.arch=="ab1r236" :
        net = ab1resnet236(h=args.h,b1=args.b1)
    elif args.arch=="ab1r242" :
        net = ab1resnet242(h=args.h,b1=args.b1)
    elif args.arch=="ab1r248" :
        net = ab1resnet248(h=args.h,b1=args.b1)
    elif args.arch=="ab1r254" :
        net = ab1resnet254(h=args.h,b1=args.b1)
    elif args.arch=="ab1r260" :
        net = ab1resnet260(h=args.h,b1=args.b1)    
    elif args.arch=="ab1r266" :
        net = ab1resnet266(h=args.h,b1=args.b1)    
        
    elif args.arch=="ab1r278" :
        net = ab1resnet278(h=args.h,b1=args.b1)
    elif args.arch=="ab1r308" :
        net = ab1resnet308(h=args.h,b1=args.b1)
    elif args.arch=="ab1r338" :
        net = ab1resnet338(h=args.h,b1=args.b1)
    elif args.arch=="ab1r368" :
        net = ab1resnet368(h=args.h,b1=args.b1)
    elif args.arch=="ab1r398" :
        net = ab1resnet398(h=args.h,b1=args.b1)
    elif args.arch=="ab1r428" :
        net = ab1resnet428(h=args.h,b1=args.b1)
    elif args.arch=="ab1r458" :
        net = ab1resnet458(h=args.h,b1=args.b1)
    elif args.arch=="ab1r488" :
        net = ab1resnet488(h=args.h,b1=args.b1)
        
    elif args.arch=="r20" :
        net = resnet20()
    elif args.arch=="r32" :
        net = resnet32()
    elif args.arch=="r44" :
        net = resnet44()
    elif args.arch=="r56" :
        net = resnet56() 
    elif args.arch=="r110" :
        net = resnet110() 
    
    elif args.arch=="abl2r20" :
        net = abl2resnet20()
    elif args.arch=="abl2r32" :
        net = abl2resnet32()
    elif args.arch=="abl2r44" :
        net = abl2resnet44()
    elif args.arch=="abl2r56" :
        net = abl2resnet56() 
    elif args.arch=="abl2r110" :
        net = abl2resnet110() 
        
    elif args.arch=="abll2r20" :
        net = abll2resnet20()
    elif args.arch=="abll2r32" :
        net = abll2resnet32()
    elif args.arch=="abll2r44" :
        net = abll2resnet44()
    elif args.arch=="abll2r56" :
        net = abll2resnet56() 
    elif args.arch=="abll2r110" :
        net = abll2resnet110() 
        
        
    elif args.arch=="rx29_2" :
        net = ResNeXt29_2x64d()    
    elif args.arch=="rx29_4" :
        net = ResNeXt29_4x64d()
    elif args.arch=="rx29_8" :
        net = ResNeXt29_8x64d()
    elif args.arch=="rx29_32" :
        net = ResNeXt29_32x4d() 
           
    net = net.to(device)
    
        
    summary(net, (3, 32, 32))
    if device == 'cuda':
        net = torch.nn.DataParallel(net)
        cudnn.benchmark = True

    if args.resume:
        # Load checkpoint.
        print('==> Resuming from checkpoint..')
        assert os.path.isdir('checkpoint'), 'Error: no checkpoint directory found!'
        checkpoint = torch.load('./checkpoint/ckpt.pth')
        net.load_state_dict(checkpoint['net'])
        best_acc = checkpoint['acc']
        start_epoch = checkpoint['epoch']

    criterion = nn.CrossEntropyLoss()
    optimizer = optim.SGD(net.parameters(), lr=args.lr, momentum=0.9, weight_decay=1e-4)
    #optimizer = SGD_atan(net.parameters(), lr=args.lr, momentum=0.9, weight_decay=5e-4, alpha=args.alpha, beta=args.beta)
    #optimizer = Adam(net.parameters(), betas=(0.9, 0.999), weight_decay=5e-4)
    #optimizer = torch.optim.RMSprop(net.parameters(), lr=args.lr, alpha=0.99, eps=1e-08, weight_decay=5e-4, momentum=0.9)
    # train_scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, args.epoch)
    # train_scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, args.epoch)
    train_scheduler = optim.lr_scheduler.MultiStepLR(optimizer, milestones=[80,120], gamma=0.1)
    # iter_per_epoch = len(trainloader)
    # warmup_scheduler = WarmUpLR(optimizer, iter_per_epoch * args.warm)

    train_time  = 0.0
    test_time = 0.0
    train_top1_acc = 0.0
    train_min_loss = 100
    test_top1_acc = 0.0
    test_min_loss = 100
    #lr_list = []
    

    writer = SummaryWriter(log_dir=args.save_path)

    tmp1213=0.0
    for epoch in range(1, args.epoch):
        if epoch > args.warm:
            train_scheduler.step(epoch)
        #lr_list.append(optimizer.param_groups[0]['lr'])
        train_acc_epoch, train_loss_epoch, train_epoch_time = train(epoch)
#        if epoch > args.warm:
#            train_scheduler.step(epoch)
        train_top1_acc = max(train_top1_acc, train_acc_epoch)
        train_min_loss = min(train_min_loss, train_loss_epoch)
        train_time += train_epoch_time
        acc, test_loss_epoch, test_epoch_time = test(epoch)
        test_top1_acc = max(test_top1_acc, acc)
        if test_top1_acc > tmp1213:
            tmp1213=test_top1_acc
            # #2步法
            # output_path='output/'+str(args.arch)+'_h_'+str(args.h)+'_b1_'+str(args.b1)+'_b2_'+str(args.b2)+'_'+str(args.repnum)+'.pth'
            
            #3步法
            # output_path='output/'+str(args.arch)+'_h_'+str(args.h)+'_b1_'+str(args.b1)+'_b2_'+str(args.b2)+'_b3_'+str(args.b3)+'_'+str(args.repnum)+'.pth'
            
            # torch.save(net, output_path)
            
        test_min_loss = min(test_min_loss, test_loss_epoch)
        test_time += test_epoch_time


    writer.close()
    end_train = train_time//60
    end_test = test_time//60
    print(model_name)
    print("train time: {}D {}H {}M".format(end_train//1440, (end_train%1440)//60, end_train%60))
    print("tset time: {}D {}H {}M".format(end_test//1440, (end_test%1440)//60, end_test%60))
    print("train_acc_top1:{}, train_min_loss:{}, train_time:{}, test_top1_acc:{}, test_min_loss:{}, test_time:{}".format(train_top1_acc, train_min_loss, train_time, test_top1_acc, test_min_loss, test_time))
    
    #2步法
    output_path='output/'+str(args.arch)+'_h_'+str(args.h)+'_b1_'+str(args.b1)+'_b2_'+str(args.b2)+'_'+str(args.repnum)+'.pth'
            
    #3步法
    # output_path='output/'+str(args.arch)+'_h_'+str(args.h)+'_b1_'+str(args.b1)+'_b2_'+str(args.b2)+'_b3_'+str(args.b3)+'_'+str(args.repnum)+'.pth'
            
    torch.save(net, output_path)
    
    rowqc=(args.tao*5)+args.nnlayer+1
    colqc=args.repnum+3
    
    from openpyxl import load_workbook
    wb = load_workbook(r'log/4.25/4.25.xlsx')
    sheet = wb.active
    sheet.cell(row=rowqc, column=colqc).value = test_top1_acc.item()
    coltem=colqc+7
    sheet.cell(row=rowqc, column=coltem).value = train_top1_acc.item()
    wb.save(r'log/4.25/4.25.xlsx')
    print("运行结束！")
    print("0.3-4.5")

    # 定义噪声类型列表和对应的噪声系数列表
    noise_types = ['randn', 'rand', 'const']
    noise_coefficients = {
        'randn': [0.01, 0.03, 0.05, 0.08, 0.1],
        'rand': [-0.2, -0.1, -0.05, 0.05, 0.1, 0.2],
        'const': [-0.4, -0.2, -0.1, 0.1, 0.2, 0.4]
    }


    from openpyxl import load_workbook
    table_loc='log/'+args.table_loc+'/'+args.table_loc+'.xlsx'
    wb = load_workbook(table_loc)
    sheet = wb.active
    rowqc=3                                             #行
    colqc=2+(args.tao*3)+(args.repnum+1)                #列
    # 嵌套循环遍历噪声类型和噪声系数
    for noise_type in noise_types:
        for noise_coefficient in noise_coefficients[noise_type]:
            noise_acc, _, _= test_noise(net,noise_type,noise_coefficient)
            print('噪声类型:'+str(noise_type)+';'+'噪声幅度'+str(noise_coefficient)+';'+'准确率'+str('%.2f' % noise_acc.item()))
            sheet.cell(row=rowqc, column=colqc).value = noise_acc.item()
            rowqc+=1
            wb.save(table_loc)